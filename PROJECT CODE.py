# Importing relevant Libraries
import pandas as pd
import numpy as np
import xlsxwriter
import math
import sys
import re
from matplotlib.pyplot import axis
from openpyxl import load_workbook
from pandas import read_excel
from os import read

sys.path.append("C:\\Users\\Susane Kiwana\\AppData\\Local\\Programs\\Python\\Python310\\Lib\\site-packages")

# Creating functions to take in parameters from the user
def age_and_term():
    age = input("Age (years): ")
    if int(age) >= 20 and int(age) <= 60:
        age_and_term = xlsxwriter.Workbook("age_and_term.xlsx")
        worksheet = age_and_term.add_worksheet("Product data")
        worksheet.write("A1", "Age (years)")  
        worksheet.write("B1", age)             

    else:
        print("Invalid age. Please enter a valid age from 20 years to 60 years.")

    term = input("Term (years): ")
    if int(term) >= 1 and int(term) <= 15 and (int(age) + int(term)) <= 75:     
        worksheet.write("A2", "Term (years)")  
        worksheet.write("B2", term)           
        age_and_term.close()
    else:
        print("Invalid term. Please enter a lower term.") 


def premium_value():
    premium = input("Single premium: ")
    if int(premium) >= 0: 
        premium_value = xlsxwriter.Workbook("premium_value.xlsx")
        worksheet = premium_value.add_worksheet("Product data")     # You cannot append to an existing xlsx file with xlsxwriter
        worksheet.write("A1", "Premium (UGX)")  
        worksheet.write("B1", premium)     
        premium_value.close()    
    else:
        print("Invalid premium. Please enter a valid charge.")


def bid_offer():
    bo_spread = input("Bid-Offer Spread: ")
    if float(bo_spread) >= 0 and float(bo_spread) <= 20:    
        bid_offer = xlsxwriter.Workbook("bid_offer.xlsx")
        worksheet = bid_offer.add_worksheet("Product data")
        worksheet.write("A1", "Bid-Offer (%)")  
        worksheet.write("B1", "{:.2%}".format(float(bo_spread)/100))   
        bid_offer.close()
    else:
        print("Invalid bid-offer spread. Please enter a valid charge.")


def interest_rate():
    interest = input("Interest rate (%): ")
    if float(interest) >= 0 and float(interest) <= 100:    
        interest_rate = xlsxwriter.Workbook("interest_rate.xlsx")
        worksheet = interest_rate.add_worksheet("Product data")
        worksheet.write("A1", "Interest rate (%)")  
        worksheet.write("B1", "{:.2%}".format(float(interest)/100))  
        interest_rate.close()  
    else:
        print("Invalid interest rate. Please enter a valid rate.")


def risk_discount():
    risk_discount_rate = input("Risk discount rate (%): ")
    if float(risk_discount_rate) >= 0 and float(risk_discount_rate) <= 100:    
        risk_discount = xlsxwriter.Workbook("risk_discount.xlsx")
        worksheet = risk_discount.add_worksheet("Product data")
        worksheet.write("A1", "Risk discount rate (%)")  
        worksheet.write("B1", "{:.2%}".format(float(risk_discount_rate)/100))  
        risk_discount.close()    
    else:
        print("Invalid risk discount rate. Please enter a valid rate.")


def allocation():
    unit_allocation_rate = input("Unit Allocation rate (%): ")
    if float(unit_allocation_rate) >= 0 and float(unit_allocation_rate) <= 100:    
        allocation = xlsxwriter.Workbook("allocation.xlsx")
        worksheet = allocation.add_worksheet("Product data")
        worksheet.write("A1", "Unit Allocation rate (%)")  
        worksheet.write("B1", "{:.2%}".format(float(unit_allocation_rate)/100)) 
    else:
        print("Invalid unit allocation rate. Please enter a valid rate.")

    non_unit_allocation_rate = input("Non-Unit Allocation rate (%): ") 
    if float(non_unit_allocation_rate) >= 0 and float(non_unit_allocation_rate) <= 100 and int(unit_allocation_rate) + int(non_unit_allocation_rate) <=100: 
        worksheet.write("A2", "Non-Unit Allocation Rate (%)")  
        worksheet.write("B2", "{:.2%}".format(float(non_unit_allocation_rate)/100))  
        allocation.close()
    else:
        print("Invalid non unit allocation rate. Please enter a valid rate.")


def unit_growth():
    growth_rate = input("Unit Fund Growth rate (%): ")
    if float(growth_rate) > 0 and float(growth_rate) <= 100:    
        unit_growth = xlsxwriter.Workbook("unit_growth.xlsx")
        worksheet = unit_growth.add_worksheet("Product data")
        worksheet.write("A1", "Unit Growth Rate (%)")  
        worksheet.write("B1", "{:.2%}".format(float(growth_rate)/100)) 
        unit_growth.close()  
    else:
        print("Invalid growth rate. Please enter a valid rate.")


def unit_tax():
    unitfund_tax_rate = input("Unit Fund Tax rate (%): ")
    if float(unitfund_tax_rate) >= 0 and float(unitfund_tax_rate) <= 100:   
        unit_tax = xlsxwriter.Workbook("unit_tax.xlsx")
        worksheet = unit_tax.add_worksheet("Product data")
        worksheet.write("A1", "Unit Tax (%)")  
        worksheet.write("B1", "{:.2%}".format(float(unitfund_tax_rate)/100)) 
        unit_tax.close()   
    else:
        print("Invalid unit fund tax rate. Please enter a valid rate.")


def non_unit_tax():
    non_unitfund_tax_rate = input("Non-Unit Fund Tax rate (%): ")
    if float(non_unitfund_tax_rate) >= 0 and float(non_unitfund_tax_rate) <= 100:   
        non_unit_tax = xlsxwriter.Workbook("non_unit_tax.xlsx")
        worksheet = non_unit_tax.add_worksheet("Product data")
        worksheet.write("A1", "Non-Unit Tax (%)")  
        worksheet.write("B1", "{:.2%}".format(float(non_unitfund_tax_rate)/100)) 
        non_unit_tax.close()     
    else:
        print("Invalid non-unit fund tax rate. Please enter a valid rate.")


def mgt_charge():
    management_charge = input("Fund Management Charge (%): ")
    if float(management_charge) >= 0 and float(management_charge) <= 20:  
        mgt_charge = xlsxwriter.Workbook("mgt_charge.xlsx")
        worksheet = mgt_charge.add_worksheet("Product data")
        worksheet.write("A1", "Management Charge (%)")  
        worksheet.write("B1", "{:.2%}".format(float(management_charge)/100))
        mgt_charge.close()     
    else:
        print("Invalid management charge. Please enter a valid charge.")


# Creating a menu to input parameters
def menu():
    print("Please Enter the Required Values Below:\n")
    premium_value()
    age_and_term()
    interest_rate()
    risk_discount()
    allocation()
    unit_growth()
    unit_tax()
    non_unit_tax()
    mgt_charge()
    bid_offer()

menu()


# Extracting parameters from Excel
mgt_wb = load_workbook("mgt_charge.xlsx")                      #this code uses openpyxl
mgt_ws = mgt_wb.active
mgt_charge = float(re.sub("%", "", mgt_ws["B1"].value)) # where input is the values that you want substituted
print("Management charge is " + str(mgt_charge) + "%")                                       # re.sub() is the function to replace the % sign with nothing

bo_spread_wb = load_workbook("bid_offer.xlsx")
bo_spread_ws = bo_spread_wb.active
bid_offer = float(re.sub("%", "", bo_spread_ws["B1"].value))
print("Bid offer spread is " + str(bid_offer) + "%" + " of Allocated Premium")

non_unit_tax_wb = load_workbook("non_unit_tax.xlsx")
non_unit_tax_ws = non_unit_tax_wb.active
non_unit_tax = float(re.sub("%", "", non_unit_tax_ws["B1"].value))
print("Non-unit tax is " + str(non_unit_tax) + "%" + " of the non-unit fund interest")

unit_tax_wb = load_workbook("unit_tax.xlsx")
unit_tax_ws = unit_tax_wb.active
unit_tax = float(re.sub("%", "", unit_tax_ws["B1"].value))
print("Unit tax is " + str(unit_tax) + "%" + " of the unit fund growth rate")

unit_growth_wb = load_workbook("unit_growth.xlsx")
unit_growth_ws = unit_growth_wb.active
unit_growth = float(re.sub("%", "", unit_growth_ws["B1"].value))
print("Unit growth rate is " + str(unit_growth) + "%")

allocation_wb = load_workbook("allocation.xlsx")
allocation_ws = allocation_wb.active
allocation = float(re.sub("%", "", allocation_ws["B1"].value))
print("Premium percentage allocated to the unit fund is " + str(allocation) + "%")

non_allocation_wb = load_workbook("allocation.xlsx")
non_allocation_ws = allocation_wb.active
non_allocation = float(re.sub("%", "", non_allocation_ws["B2"].value))
print("Premium percentage allocated to the non-unit fund is " + str(non_allocation) + "%")

interest_wb = load_workbook("interest_rate.xlsx")
interest_ws = interest_wb.active
interest_rate = float(re.sub("%", "", interest_ws["B1"].value))
print("Interest rate on the non-unit fund is " + str(interest_rate) + "%")

risk_discount_wb = load_workbook("risk_discount.xlsx")
risk_discount_ws = risk_discount_wb.active
risk_discount = float(re.sub("%", "", risk_discount_ws["B1"].value))
print("Risk discount rate is " + str(risk_discount) + "%")

premium_value_wb = load_workbook("premium_value.xlsx")       # if i wanted to pull out 2 cells, i could use the f string as so: 
premium_value_ws = premium_value_wb.active                   # premium_value = f"{premium_value_ws["A1"].value}: {premium_value_ws["B1"].value}"
premium_value = int(premium_value_ws["B1"].value)            # OUTPUT BEING: Premium value: 20000
print("Premium value is " + str(premium_value) + " UGX")

age_wb = load_workbook("age_and_term.xlsx")
age_ws = age_wb.active            
age = int(age_ws["B1"].value)
print("Age of the policy holder is " + str(age) + " years")

term_wb = load_workbook("age_and_term.xlsx")
term_ws = term_wb.active            
term = int(term_ws["B2"].value)
print("Term of the policy is " + str(term) + " years")


# Importing Multiple decrement tables
male_mdt = pd.read_excel("TABLE.xlsx","Mortality Data (Males)", skiprows = range(1,age - 19), skipfooter = 101 - (age + term)) # keep the first row 0 (as the header) and then skip everything else up to policy holder age and then skip last bottom rows beyond term
male_mdt 

female_mdt = pd.read_excel("TABLE.xlsx","Mortality Data (Females)", skiprows = range(1,age - 19), skipfooter = 101 - (age + term))
female_mdt


# Creating columns for Age and Term
year = term
age_col = pd.DataFrame(columns=["Age"])       # I want to put the ouput of age and term in a dataframe. 
year_col = pd.DataFrame(columns=["Year"])     # I've decided to create to separate tables with the hope of merging them later.

if int(year) >= 1 and int(year) <= 15 and (int(age) + int(year)) <= 75:       
    for year in range(1, int(year)+1):
        year_col.loc[year, "Year"] = year                       #trying to append term to my 2nd dataframe as column 1
    
else:
    print("Invalid term. Please enter a lower term.")

if int(age) >= 20 and int(age) <= 60: 
    for age in range(int(age), int(age) + int(year)):
        age_col.loc[age, "Age"] = age         # trying to append age to my dataframe as column 1. The loc function indexes, e.g. df.loc(row,column) = randint(0,99)

else:
    print("Invalid age. Please enter a valid age from 20 years to 100 years.")      
    

age_col2 = age_col.reset_index(drop = True)           # Trying to make the index start from 0 instead of from the number I input
year_col2 = year_col.reset_index(drop = True)

age_year_cols = pd.concat([age_col2,year_col2], axis = 1)         #axis makes everything start from the beginning

age_year_cols.to_excel("age_and_term.xlsx",index = False)



# Creating other columns for unit fund table
unit_data = age_year_cols
   
unit_data["Fund value at the start of the Year"] = 0

unit_data["Allocated Premium"] = (allocation/100) * premium_value

unit_data["B/O Spread"] = (bid_offer/100) * unit_data["Allocated Premium"]


for ind, row in unit_data.iterrows():
    unit_data.iloc[1:,3] = 0    #setting the zeros in column allocated premium
    
    unit_data.iloc[1:,4] = 0    #setting the zeros in column b/o spread

    unit_data["Interest"] = (unit_growth/100) * (unit_data["Fund value at the start of the Year"] + unit_data["Allocated Premium"] - unit_data["B/O Spread"])
        
    unit_data["Unit Fund Tax"] = (unit_tax/100) * (unit_data["Interest"])

    unit_data["Fund Management Charge"] = (mgt_charge/100) * (unit_data["Fund value at the start of the Year"] + unit_data["Allocated Premium"] - unit_data["B/O Spread"] + unit_data["Interest"] - unit_data["Unit Fund Tax"])

    unit_data["Fund value at the end of the Year"] = unit_data["Fund value at the start of the Year"] + unit_data["Allocated Premium"] - unit_data["B/O Spread"] + unit_data["Interest"] - unit_data["Unit Fund Tax"] - unit_data["Fund Management Charge"]

    unit_data.iloc[1:,2] = unit_data.iloc[0:unit_data.shape[0]-1,8]    #setting the fund value at the start of the year. unit_data.shape[0]-1 is the 2nd last row
 

unit_data.round()



# Creating columns for non-unit Fund table (Males)
non_unit_data_males = pd.read_excel("age_and_term.xlsx")

non_unit_data_males["Unallocated Premium"] = (non_allocation/100) * premium_value

non_unit_data_males["B/O Spread"] = (bid_offer/100) * unit_data["Allocated Premium"]

non_unit_data_males["Expenses"] = 0.05 * premium_value

non_unit_data_males["Commission"] = 0.02 * premium_value

for ind, row in non_unit_data_males.iterrows():
    non_unit_data_males.iloc[1:,2] = 0    #setting the zeros in column unallocated premium
    
    non_unit_data_males.iloc[1:,4] = 0    #setting the zeros in column expenses
    
    non_unit_data_males.iloc[1:,5] = 0    #setting the zeros in column commission

    non_unit_data_males["Interest"] = (interest_rate/100) * (non_unit_data_males["Unallocated Premium"] + unit_data["B/O Spread"] - non_unit_data_males["Expenses"] - non_unit_data_males["Commission"])

    non_unit_data_males["Non-Unit Fund Tax"] = (non_unit_tax/100) * (non_unit_data_males["Interest"])

    non_unit_data_males["Fund Management Charge"] = ((mgt_charge/100) * (unit_data["Fund value at the start of the Year"] + unit_data["Allocated Premium"] - unit_data["B/O Spread"] + unit_data["Interest"] - unit_data["Unit Fund Tax"])).round()

    non_unit_data_males["Extra Death Cost"] = np.where((male_mdt["(aq)x (d)"]*(premium_value - unit_data["Fund value at the end of the Year"])) > 0, (male_mdt["(aq)x (d)"]*(premium_value - unit_data["Fund value at the end of the Year"])).round(), 0)   #here the numpy.where is my if condition for the max value expected death benefit

    non_unit_data_males["Extra Surrender Cost"] = (male_mdt["(aq)x (w)"] * (0.7 * (premium_value - unit_data["Fund value at the end of the Year"]))).round()

    non_unit_data_males["Expected Claim Expenses"] = ((male_mdt["(aq)x (d)"] + male_mdt["(aq)x (w)"]) * (0.07 * premium_value)).round()

    non_unit_data_males.iloc[-1,10] = 0    #setting the last row in column extra surrender cost to zero
    
    non_unit_data_males.iloc[-1,11] = (male_mdt.iloc[-1,3] + male_mdt.iloc[-1,4] + male_mdt.iloc[-1,5]) * (0.07 * premium_value)    #setting the last row in column extra surrender cost to zero

    non_unit_data_males["End of year cashflows"] = (non_unit_data_males["Unallocated Premium"] + non_unit_data_males["B/O Spread"] - non_unit_data_males["Expenses"] - non_unit_data_males["Commission"] + non_unit_data_males["Interest"] - non_unit_data_males["Non-Unit Fund Tax"] + non_unit_data_males["Fund Management Charge"] - non_unit_data_males["Extra Death Cost"] - non_unit_data_males["Extra Surrender Cost"] - non_unit_data_males["Expected Claim Expenses"]).round()

    non_unit_data_males["Probability in force at the start of the year"] = male_mdt["t-1(ap)x"]

    non_unit_data_males["Profit Signature"] = (non_unit_data_males["End of year cashflows"] * non_unit_data_males["Probability in force at the start of the year"]).round()

    non_unit_data_males["Discount factor"] = (1+(risk_discount/100)) ** (non_unit_data_males["Year"])

    non_unit_data_males["EPV of Profit"] = (non_unit_data_males["Profit Signature"] * non_unit_data_males["Discount factor"]).round()

non_unit_data_males



# Creating columns for non-unit Fund table (Females)
non_unit_data_females = pd.read_excel("age_and_term.xlsx")

non_unit_data_females["Unallocated Premium"] = (non_allocation/100) * premium_value

non_unit_data_females["B/O Spread"] = (bid_offer/100) * unit_data["Allocated Premium"]

non_unit_data_females["Expenses"] = 0.05 * premium_value

non_unit_data_females["Commission"] = 0.02 * premium_value

for ind, row in non_unit_data_females.iterrows():
    non_unit_data_females.iloc[1:,2] = 0    #setting the zeros in column unallocated premium
    
    non_unit_data_females.iloc[1:,4] = 0    #setting the zeros in column expenses
    
    non_unit_data_females.iloc[1:,5] = 0    #setting the zeros in column commission

    non_unit_data_females["Interest"] = (interest_rate/100) * (non_unit_data_females["Unallocated Premium"] + unit_data["B/O Spread"] - non_unit_data_females["Expenses"] - non_unit_data_females["Commission"])

    non_unit_data_females["Non-Unit Fund Tax"] = (non_unit_tax/100) * (non_unit_data_females["Interest"])

    non_unit_data_females["Fund Management Charge"] = ((mgt_charge/100) * (unit_data["Fund value at the start of the Year"] + unit_data["Allocated Premium"] - unit_data["B/O Spread"] + unit_data["Interest"] - unit_data["Unit Fund Tax"])).round()

    non_unit_data_females["Extra Death Cost"] = np.where((female_mdt["(aq)x (d)"]*(premium_value - unit_data["Fund value at the end of the Year"])) > 0, (female_mdt["(aq)x (d)"]*(premium_value - unit_data["Fund value at the end of the Year"])).round(), 0)   #here the numpy.where is my if condition for the max value expected death benefit

    non_unit_data_females["Extra Surrender Cost"] = (female_mdt["(aq)x (w)"] * (0.7 * (premium_value - unit_data["Fund value at the end of the Year"]))).round()

    non_unit_data_females["Expected Claim Expenses"] = ((female_mdt["(aq)x (d)"] + female_mdt["(aq)x (w)"]) * (0.07 * premium_value)).round()

    non_unit_data_females.iloc[-1,10] = 0    #setting the last row in column extra surrender cost to zero
    non_unit_data_females.iloc[-1,11] = (female_mdt.iloc[-1,3] + female_mdt.iloc[-1,4] + female_mdt.iloc[-1,5]) * (0.07 * premium_value)    #setting the last row in column extra surrender cost to zero

    non_unit_data_females["End of year cashflows"] = (non_unit_data_females["Unallocated Premium"] + non_unit_data_females["B/O Spread"] - non_unit_data_females["Expenses"] - non_unit_data_females["Commission"] + non_unit_data_females["Interest"] - non_unit_data_females["Non-Unit Fund Tax"] + non_unit_data_females["Fund Management Charge"] - non_unit_data_females["Extra Death Cost"] - non_unit_data_females["Extra Surrender Cost"] - non_unit_data_females["Expected Claim Expenses"]).round()

    non_unit_data_females["Probability in force at the start of the year"] = female_mdt["t-1(ap)x"]

    non_unit_data_females["Profit Signature"] = (non_unit_data_females["End of year cashflows"] * non_unit_data_females["Probability in force at the start of the year"]).round()

    non_unit_data_females["Discount factor"] = (1+(risk_discount/100)) ** (non_unit_data_females["Year"])

    non_unit_data_females["EPV of Profit"] = (non_unit_data_females["Profit Signature"] * non_unit_data_females["Discount factor"]).round()


with pd.ExcelWriter("Model.xlsx") as writer:                                        #adding a third sheet to the file Model
    unit_data.round().to_excel(writer,sheet_name = "UNIT FUND", index = False)
    non_unit_data_males.to_excel(writer, sheet_name = "NON-UNIT FUND (Males)", index = False)
    non_unit_data_females.to_excel(writer, sheet_name = "NON-UNIT FUND (Females)", index = False)

non_unit_data_females


NPV_males = non_unit_data_males["EPV of Profit"].sum()
Profit_Margin_males = (NPV_males/premium_value)

Profit_Metrics = xlsxwriter.Workbook("Profit Metrics.xlsx")
worksheet = Profit_Metrics.add_worksheet("Males")
worksheet.write("A1", "NPV")  
worksheet.write("B1", NPV_males)
worksheet.write("A2", "Profit Margin (%)")  
worksheet.write("B2", "{:.4%}".format(float(Profit_Margin_males))) 

NPV_females = non_unit_data_females["EPV of Profit"].sum()
Profit_Margin_females = (NPV_females/premium_value)

worksheet = Profit_Metrics.add_worksheet("Females")
worksheet.write("A1", "NPV")  
worksheet.write("B1", NPV_females)
worksheet.write("A2", "Profit Margin (%)")  
worksheet.write("B2", "{:.4%}".format(float(Profit_Margin_females))) 

Profit_Metrics.close()  